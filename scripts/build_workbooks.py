"""
Build the two delivery workbooks from validated CSVs.

Design intent: a reader must be able to tell a reported number from a modeled
one at a glance, without reading documentation. Modeled cells are tinted;
reported cells are not. The README tab states the methodology in plain language
before the reader reaches a single number.
"""
import re
import sys
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

INK = "28251D"
TEAL = "01696F"
BAND = "F2F0EB"
MODELED_TINT = "FFF6E8"     # warm tint = modeled
REPORTED_TINT = "EAF3F1"    # cool tint = reported
BORDER = "D4D1CA"

HEAD_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor=TEAL)
BODY_FONT = Font(name="Calibri", size=10, color=INK)
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SCENT_HINTS = [
    "Mint", "Peppermint", "Spearmint", "Wintergreen", "Cinnamon", "Clean Mint",
    "Cool Mint", "Fresh Mint", "Charcoal", "Coconut", "Lavender", "Citrus",
    "Lemon", "Aloe", "Original", "Unscented", "Fragrance Free", "Ocean",
    "Vanilla", "Strawberry", "Bubble Fruit", "Watermelon", "Shea Butter",
    "Milk & Honey", "Cucumber", "Apricot", "Rosemary", "Eucalyptus",
]

FORM_MAP = [
    (r"toothpaste|dental cream|gel", "Paste / Gel"),
    (r"toothbrush|brush head", "Brush"),
    (r"mouthwash|rinse|mouth rinse", "Liquid Rinse"),
    (r"floss|pick|interdental", "Floss / Pick"),
    (r"strip|pen|whitening kit|led", "Whitening Device / Strip"),
    (r"bar soap|bar\b", "Bar"),
    (r"body wash|hand soap|shower gel|cleanser|liquid", "Liquid"),
    (r"foaming", "Foam"),
    (r"antiperspirant|deodorant|stick", "Stick / Solid"),
    (r"spray|mist", "Spray"),
    (r"serum|treatment|peel", "Serum / Treatment"),
    (r"cream|lotion|moisturizer|balm", "Cream / Lotion"),
    (r"sunscreen|spf", "Sunscreen"),
    (r"wipe", "Wipe"),
]


def derive_variant(name: str) -> str:
    """Pull a scent/flavour descriptor out of the product name.

    Conservative by design — returns 'n.a.' rather than guessing, because a
    wrong variant silently splits what should be one product into two.
    """
    for s in sorted(SCENT_HINTS, key=len, reverse=True):
        if re.search(rf"\b{re.escape(s)}\b", name, re.I):
            return s
    return "n.a."


def derive_form(name: str, sub_category: str) -> str:
    hay = f"{sub_category} {name}".lower()
    for pat, form in FORM_MAP:
        if re.search(pat, hay):
            return form
    return "n.a."


def slug(s: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(s).upper())[:12] or "NA"


def style_sheet(ws, df, freeze="A2", tint_cols=None, tint_by_source=False):
    tint_cols = tint_cols or []
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill, cell.border = HEAD_FONT, HEAD_FILL, BOX
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    src_idx = list(df.columns).index("source_type") + 1 if "source_type" in df.columns else None

    for r, row in enumerate(df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font, cell.border = BODY_FONT, BOX
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if tint_by_source and src_idx:
            st = str(ws.cell(row=r, column=src_idx).value or "")
            for c in tint_cols:
                tint = MODELED_TINT if st in ("Modeled", "Mixed") else REPORTED_TINT
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=tint)
        elif r % 2 == 0:
            for c in range(1, len(df.columns) + 1):
                if not ws.cell(row=r, column=c).fill.fgColor.rgb or \
                        ws.cell(row=r, column=c).fill.patternType is None:
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BAND)

    for c, name in enumerate(df.columns, start=1):
        longest = max([len(str(name))] + [len(str(v)) for v in df.iloc[:, c - 1].head(300)])
        ws.column_dimensions[get_column_letter(c)].width = min(max(12, longest + 2), 52)

    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"


def readme_tab(wb, title, lines):
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 104
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=TEAL)
    r = 3
    for kind, text in lines:
        c = ws.cell(row=r, column=1, value=text)
        if kind == "h":
            c.font = Font(name="Calibri", bold=True, size=11, color=INK)
            r += 1
        else:
            c.font = Font(name="Calibri", size=10, color=INK)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(15, 14 * (len(text) // 100 + 1))
            r += 1
    return ws
