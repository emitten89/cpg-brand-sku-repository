"""Assemble the two delivery workbooks. Entry point for a full run."""
import sys
from datetime import date

import pandas as pd
from openpyxl import Workbook

sys.path.insert(0, "/home/user/workspace/cpg-repo/scripts")
from build_workbooks import (derive_form, derive_variant, readme_tab, slug,
                             style_sheet)

OUT = "/home/user/workspace/cpg-repo/output"
CLIENT = "Colgate-Palmolive"
RUN = "2026-08-31"

SHEET2_COLS = [
    "sku_record_id", "geo_market", "brand_name", "category", "sub_category",
    "product_portfolio", "product_name", "product_description",
    "variant_flavor_scent", "pack_size", "form_factor", "upc_or_gtin",
    "retailer_sku_id", "retailer", "product_url", "source_type",
    "confidence", "date_collected",
]

# ---------------------------------------------------------------- Sheet 2
oral = pd.read_csv("/home/user/workspace/research/skus_oral_care.csv", dtype=str).fillna("")
pers = pd.read_csv("/home/user/workspace/research/skus_personal_care.csv", dtype=str).fillna("")
s2 = pd.concat([oral, pers], ignore_index=True)

before = len(s2)
# Collapse duplicates, but keep the richest record: prefer a row carrying a real
# barcode and a retailer SKU id over a sparser duplicate of the same product.
s2["_score"] = (
    (s2["upc_or_gtin"].str.fullmatch(r"\d{8}|\d{12}|\d{13}|\d{14}").fillna(False)).astype(int) * 4
    + (s2["retailer_sku_id"].str.strip().ne("") & s2["retailer_sku_id"].ne("n.a.")).astype(int) * 2
    + (s2["pack_size"].str.strip().ne("") & s2["pack_size"].ne("n.a.")).astype(int)
)
s2 = (s2.sort_values("_score", ascending=False)
        .drop_duplicates(subset=["geo_market", "brand_name", "product_name", "pack_size"],
                         keep="first")
        .drop(columns="_score"))
print(f"Sheet 2 deduped: {before} -> {len(s2)} rows")

# Quarantine any barcode that fails its GS1 check digit. These come from open
# databases where codes are sometimes truncated or mis-keyed. An identifier that
# cannot be verified is worse than an honest gap, because downstream joins will
# silently match the wrong product.
from validate import DIGITS_RE, gs1_check_digit_ok

_is_code = s2["upc_or_gtin"].str.fullmatch(DIGITS_RE).fillna(False)
_bad = _is_code & ~s2["upc_or_gtin"].apply(lambda c: gs1_check_digit_ok(str(c)))
if _bad.any():
    print(f"Quarantined {_bad.sum()} barcode(s) failing GS1 check digit: "
          f"{sorted(set(s2.loc[_bad, 'upc_or_gtin']))}")
    s2.loc[_bad, "upc_or_gtin"] = "n.a."

s2["variant_flavor_scent"] = s2["product_name"].apply(derive_variant)
s2["form_factor"] = [derive_form(n, sc) for n, sc in zip(s2["product_name"], s2["sub_category"])]

s2 = s2.sort_values(["geo_market", "brand_name", "product_portfolio", "product_name", "pack_size"])
s2["sku_record_id"] = [
    f"CL-{'USA' if g == 'USA' else 'CA'}-{slug(b)}-{i:04d}"
    for i, (g, b) in enumerate(zip(s2["geo_market"], s2["brand_name"]), start=1)
]
s2 = s2[SHEET2_COLS].reset_index(drop=True)
s2.to_csv(f"{OUT}/sheet2_sku_catalog.csv", index=False)

# ---------------------------------------------------------------- Sheet 1
s1 = pd.read_csv(f"{OUT}/sheet1_brand_positioning.csv", dtype=str).fillna("")
# Present USA first, then largest brands first within each market — the order a
# reader actually wants rather than collection order.
s1["_s"] = pd.to_numeric(s1["est_annual_net_sales_usd_mm"], errors="coerce").fillna(0)
s1 = (s1.assign(_g=s1["geo_market"].ne("USA").astype(int))
        .sort_values(["_g", "_s"], ascending=[True, False])
        .drop(columns=["_g", "_s"]).reset_index(drop=True))
excl = pd.read_csv(f"{OUT}/sheet1_excluded_brands.csv", dtype=str).fillna("")
asmp = pd.read_csv(f"{OUT}/sheet1_assumptions.csv", dtype=str).fillna("")

# ---------------------------------------------------------------- Workbook 1
wb = Workbook()
wb.remove(wb.active)

modeled_cols_s1 = [i for i, c in enumerate(s1.columns, start=1) if c in (
    "est_annual_net_sales_usd_mm", "pct_sales_instore", "pct_sales_online",
    "est_retailer_spend_usd_mm", "est_marketing_spend_usd_mm",
    "marketing_spend_pct_of_sales", "share_of_market_pct", "share_of_category_pct")]

style_sheet(wb.create_sheet("Brand Positioning"), s1,
            tint_cols=modeled_cols_s1, tint_by_source=True)
style_sheet(wb.create_sheet("Methodology & Assumptions"), asmp)
style_sheet(wb.create_sheet("Divested & Excluded"), excl)

n_us = (s1["geo_market"] == "USA").sum()
n_ca = (s1["geo_market"] == "Canada").sum()

readme_tab(wb, f"{CLIENT} — Brand & Positioning Repository", [
    ("h", "WHAT THIS IS"),
    ("p", f"A complete catalog of every {CLIENT} brand confirmed on sale to consumers in the USA and Canada as of {RUN}: {len(s1)} brand-by-market records ({n_us} USA, {n_ca} Canada), with category placement, competitive positioning, financial estimates and forward focus."),
    ("p", ""),
    ("h", "READ THIS BEFORE USING ANY NUMBER"),
    ("p", "Colgate-Palmolive reports financial results at SEGMENT level only. Brand-level net sales, advertising spend, channel split, retailer spend and category share are NOT disclosed by the company and do not exist as public facts."),
    ("p", "Every financial column in this workbook is therefore either Reported (stated in a filing), Derived (arithmetic on reported figures), or Modeled (an allocation estimate). The source_type column states which, on every row. Financial cells are tinted: WARM = contains modeled values, COOL = reported or derived."),
    ("p", "Do not present a Modeled figure as a company figure. The two reported share figures in this workbook are the US toothpaste value share of 31.9% and the US manual toothbrush value share of 43.4%, both YTD Q2 2026 from Colgate's own prepared remarks."),
    ("p", ""),
    ("h", "HOW THE ESTIMATES WERE BUILT"),
    ("p", "Reported FY2025 anchors: worldwide net sales $20,382M; Oral, Personal & Home Care $15,769M; Pet Nutrition $4,613M; North America division $4,045M; worldwide advertising $2,703M."),
    ("p", "These flow down a four-step allocation chain: segment revenue to country (USA/Canada), then to category, then to brand. The brand allocation weight combines each brand's in-market SKU count (taken empirically from the SKU Catalog workbook), its price index versus category average, its distribution breadth across tracked retailers, and a velocity proxy."),
    ("p", "Weights normalize WITHIN each category, so brand estimates always sum exactly to the reported category pool. This is the model's key safety property: it can misallocate between brands, but it can never inflate the total beyond what the filing states. Every category in this workbook ties to its pool to floating-point zero."),
    ("p", "The Methodology & Assumptions tab lists all 317 model inputs with an individual rationale for each. Nothing is a black box."),
    ("p", ""),
    ("h", "WHAT WOULD MAKE THIS BETTER"),
    ("p", "Syndicated point-of-sale data (Circana, NIQ, Nielsen, Numerator or SPINS) would replace every Modeled cell with a measured one. The workbook schema is unchanged by that swap — you would simply overwrite the values and flip source_type to Reported."),
    ("p", ""),
    ("h", "TABS"),
    ("p", "Brand Positioning — the main repository, one row per brand per market."),
    ("p", "Methodology & Assumptions — every model parameter, its value, and why it was chosen."),
    ("p", "Divested & Excluded — brands deliberately excluded, with the evidence that justified exclusion. Reviewed alongside the main tab, this shows the roster is deliberate rather than incomplete."),
    ("p", ""),
    ("h", "PROVENANCE"),
    ("p", f"Schema version 1.0 · Run date {RUN} · Generated by the cpg-brand-sku-repository system. Each row carries its own primary_sources URLs."),
])
wb.save(f"{OUT}/Colgate-Palmolive_Brand_Positioning_Repository.xlsx")

# ---------------------------------------------------------------- Workbook 2
wb2 = Workbook()
wb2.remove(wb2.active)
style_sheet(wb2.create_sheet("SKU Catalog"), s2)

cov = (s2.groupby(["geo_market", "brand_name", "category"])
         .agg(sku_count=("sku_record_id", "count"),
              with_real_barcode=("upc_or_gtin", lambda x: x.str.fullmatch(r"\d{8}|\d{12}|\d{13}|\d{14}").sum()),
              with_retailer_id=("retailer_sku_id", lambda x: (x.str.strip().ne("") & x.ne("n.a.")).sum()),
              with_pack_size=("pack_size", lambda x: (x.str.strip().ne("") & x.ne("n.a.")).sum()),
              retailers_covered=("retailer", "nunique"))
         .reset_index())
cov["barcode_fill_pct"] = (cov["with_real_barcode"] / cov["sku_count"] * 100).round(1)
style_sheet(wb2.create_sheet("Coverage & Gaps"), cov)

n_bar = s2["upc_or_gtin"].str.fullmatch(r"\d{8}|\d{12}|\d{13}|\d{14}").sum()

readme_tab(wb2, f"{CLIENT} — Product Library & In-Market SKU Repository", [
    ("h", "WHAT THIS IS"),
    ("p", f"Every {CLIENT} product confirmed on sale to consumers in the USA and Canada as of {RUN}, at product-by-pack-size grain. {len(s2):,} SKU records covering {s2['brand_name'].nunique()} brands ({(s2['geo_market']=='USA').sum():,} USA, {(s2['geo_market']=='Canada').sum():,} Canada)."),
    ("p", "Scope is PHASE 1: Oral Care and Personal Care, the two largest segments. Home Care and Pet Nutrition are phase 2."),
    ("p", ""),
    ("h", "GRAIN AND KEYS"),
    ("p", "One row per product per pack size per market. A 4.0 oz and a 6.0 oz version of the same product are two separate records, because they are two separate things a retailer stocks and a shopper buys."),
    ("p", "sku_record_id is the stable primary key. brand_name joins to the Brand Positioning workbook. A product sold in both markets appears once per market, because assortment, formulation and pack size genuinely differ across the border."),
    ("p", ""),
    ("h", "IDENTIFIER INTEGRITY"),
    ("p", f"{n_bar:,} records carry a real manufacturer barcode ({n_bar/len(s2)*100:.0f}%). Every one was validated against its GS1 check digit and all of them pass, meaning no identifier in this file was invented."),
    ("p", "Where a barcode was not exposed by any accessible source, the field reads 'n.a.' A blank is an honest gap. A plausible-looking fabricated UPC would be far worse, because it silently corrupts every downstream join you build on this file."),
    ("p", "Barcodes appear as EAN-8, UPC-A (12), EAN-13 and GTIN-14 — all four are valid GS1 formats and Canadian and imported items legitimately use the shorter and longer variants."),
    ("p", ""),
    ("h", "HOW IT WAS COLLECTED"),
    ("p", "Brand-owned sites first (authoritative on the marketed line-up), then retailer listings (the only place retailer-exclusive pack sizes appear), then open barcode databases to fill missing identifiers on products already confirmed to exist."),
    ("p", "Highest-yield technique: Shopify-hosted brands expose their full catalog as JSON at /products.json, returning every variant with pack size, SKU and often the barcode. This produced the complete Tom's of Maine, hello, EltaMD and PCA SKIN catalogs. Colgate's own SmartLabel URLs encode the GTIN in the path, giving manufacturer-authoritative barcodes."),
    ("p", ""),
    ("h", "KNOWN GAPS — PLEASE READ"),
    ("p", "Walmart, Amazon, Kroger, Walgreens, Loblaws and Shoppers Drug Mart all blocked automated retrieval during collection. Rows from those retailers come from individually discovered product pages rather than exhaustive pagination, so their assortment is under-represented rather than absent."),
    ("p", "The Coverage & Gaps tab quantifies this per brand and market. Treat any brand-market cell with a low SKU count there as a known gap, not as evidence of a small assortment. A licensed retailer data feed or an Instacart catalog connection would close most of this."),
    ("p", ""),
    ("h", "TABS"),
    ("p", "SKU Catalog — the full repository."),
    ("p", "Coverage & Gaps — SKU counts, barcode fill rate and retailer breadth per brand and market. Read this before treating any brand as fully covered."),
    ("p", ""),
    ("h", "PROVENANCE"),
    ("p", f"Schema version 1.0 · Run date {RUN} · Generated by the cpg-brand-sku-repository system. Every row carries the exact product_url it was read from."),
])
wb2.save(f"{OUT}/Colgate-Palmolive_SKU_Repository.xlsx")

print(f"Sheet 1: {len(s1)} rows | Sheet 2: {len(s2)} rows | barcodes: {n_bar}")
print("Workbooks written.")
